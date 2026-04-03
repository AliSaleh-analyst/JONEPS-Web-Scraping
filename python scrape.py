import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


def load_existing_tender_nos(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                existing.add(str(row[0]).strip())
        return existing
    except FileNotFoundError:
        return set()


def save_to_excel(data_list, filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Tender_no", "Title", "Closing_date"])

    for row in data_list:
        ws.append(list(row))

    wb.save(filename)


def find_in_all_contexts(driver, find_fn):
    driver.switch_to.default_content()
    result = find_fn(driver)
    if result is not None:
        return result

    frames = driver.find_elements(By.CSS_SELECTOR, "frame, iframe")
    for i in range(len(frames)):
        driver.switch_to.default_content()
        try:
            driver.switch_to.frame(i)
            result = find_fn(driver)
            if result is not None:
                return result
            inner_frames = driver.find_elements(By.CSS_SELECTOR, "frame, iframe")
            for j in range(len(inner_frames)):
                try:
                    driver.switch_to.frame(j)
                    result = find_fn(driver)
                    if result is not None:
                        return result
                except:
                    pass
                driver.switch_to.parent_frame()
        except:
            pass
    return None


def scrape_joneps(existing_nos):
    chrome_options = Options()
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 20)
    scraped_data = []

    try:
        print("1. Opening JONEPS Main Page...")
        driver.get("https://www.joneps.gov.jo/pt/main.do")
        time.sleep(8)

        found_menu = False
        print("Searching for Tender Invitation link...")

        for attempt in range(3):
            all_frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")

            driver.switch_to.default_content()
            for text in ["Tender Invitation", "Invitations", "دعوات"]:
                try:
                    btn = driver.find_element(By.PARTIAL_LINK_TEXT, text)
                    driver.execute_script("arguments[0].click();", btn)
                    found_menu = True
                    break
                except:
                    continue

            if found_menu: break

            for i in range(len(all_frames)):
                driver.switch_to.default_content()
                try:
                    driver.switch_to.frame(i)
                    for text in ["Tender Invitation", "Invitations", "دعوات"]:
                        try:
                            btn = driver.find_element(By.PARTIAL_LINK_TEXT, text)
                            print(f"Found menu in frame {i}. Clicking...")
                            driver.execute_script("arguments[0].click();", btn)
                            found_menu = True
                            break
                        except: continue
                    if found_menu: break
                except: continue

            if found_menu: break
            print(f"Attempt {attempt+1} failed to find menu, retrying...")
            time.sleep(3)

        if not found_menu:
            time.sleep(5)
            if "main.do" not in driver.current_url:
                found_menu = True

        if not found_menu:
            print("Error: Could not find the Tender Invitation link.")
            return []

        print("Menu clicked. Waiting 20 seconds for page to fully load...")
        time.sleep(20)

        tender_links = []
        for attempt in range(10):
            tender_links = driver.find_elements(By.XPATH, "//a[contains(text(), '202') and contains(text(), '-')]")
            if tender_links:
                print(f"Table found in current frame context after ~{attempt * 3} seconds.")
                break

            driver.switch_to.default_content()
            tender_links = driver.find_elements(By.XPATH, "//a[contains(text(), '202') and contains(text(), '-')]")
            if tender_links:
                print(f"Table found in default context after ~{attempt * 3} seconds.")
                break

            all_frames = driver.find_elements(By.CSS_SELECTOR, "frame, iframe")
            for i in range(len(all_frames)):
                driver.switch_to.default_content()
                try:
                    driver.switch_to.frame(i)
                    tender_links = driver.find_elements(By.XPATH, "//a[contains(text(), '202') and contains(text(), '-')]")
                    if tender_links:
                        print(f"Table found in frame[{i}] after ~{attempt * 3} seconds.")
                        break
                    inner = driver.find_elements(By.CSS_SELECTOR, "frame, iframe")
                    for j in range(len(inner)):
                        try:
                            driver.switch_to.frame(j)
                            tender_links = driver.find_elements(By.XPATH, "//a[contains(text(), '202') and contains(text(), '-')]")
                            if tender_links:
                                print(f"Table found in frame[{i}][{j}] after ~{attempt * 3} seconds.")
                                break
                        except:
                            pass
                        driver.switch_to.parent_frame()
                    if tender_links:
                        break
                except:
                    continue
            if tender_links:
                break

            print(f"  Waiting... attempt {attempt + 1}/10")
            time.sleep(3)

        if not tender_links:
            print("Error: Could not locate the results table after 30 seconds.")
            print("Current URL:", driver.current_url)
            all_links = driver.find_elements(By.TAG_NAME, "a")
            print(f"Total links on page: {len(all_links)}")
            for a in all_links[:20]:
                print(f"  '{a.text.strip()[:60]}' -> {(a.get_attribute('href') or '')[:80]}")
            return []

        page_num = 1
        while True:
            print(f"--- Scraping page {page_num} ---")

            if page_num > 1:
                tender_links = driver.find_elements(By.XPATH, "//a[contains(text(), '202') and contains(text(), '-')]")
            if not tender_links:
                print(f"No rows found on page {page_num}, stopping.")
                break

            print(f"Found {len(tender_links)} tender links. Extracting rows...")
            new_on_page = 0
            for link in tender_links:
                try:
                    row = link.find_element(By.XPATH, "./ancestor::tr[1]")
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) < 6:
                        continue

                    t_no    = cols[0].text.strip()
                    t_title = cols[1].text.strip()
                    t_date  = cols[5].text.strip()

                    if t_no:
                        if t_no in existing_nos:
                            continue
                        scraped_data.append((t_no, t_title, t_date))
                        new_on_page += 1
                        safe_title = t_title[:40].encode("ascii", "replace").decode("ascii")
                        print(f"  Scraped: {t_no} | {safe_title}")
                except Exception as e:
                    print(f"  Row error: {e}")
                    continue

            if new_on_page == 0:
                print(f"No new tenders found on page {page_num}. Stopping pagination.")
                break

            next_btn = driver.find_elements(By.CSS_SELECTOR, "a.btn_p_next")
            if not next_btn or page_num >= 60:
                print("All pages scraped.")
                break

            page_num += 1
            print(f"Navigating to page {page_num}...")
            driver.execute_script(f"fn_pageview({page_num});")
            time.sleep(4)

    except Exception as e:
        print(f"Scraping Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()

    return scraped_data


if __name__ == "__main__":
    print("--- STARTING JONEPS AUTOMATION ---")
    FILENAME = "JONEPS-Tenders.xlsx"
    existing_nos = load_existing_tender_nos(FILENAME)
    print(f"Loaded {len(existing_nos)} existing tender numbers from {FILENAME}.")

    results = scrape_joneps(existing_nos)

    if results:
        save_to_excel(results, FILENAME)
        print(f"\n{len(results)} new tenders added to {FILENAME}.")
    else:
        print("No new tenders were collected.")
